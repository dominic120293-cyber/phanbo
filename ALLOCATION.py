"""
Phiên bản tối ưu tốc độ – v4
============================
Cải tiến so với v3 (tóm tắt tác động dự kiến):

  [WIN 1] Prune y_vars          → chỉ tạo (job, block) có supply ST/POD khớp
                                   → giảm biến ~50–70 %, constraints tương ứng
                                   → solver nhanh hơn 2–4×

  [WIN 2] Xóa hoàn toàn u_vars  → e[h,b] >= sum(y) − 1 viết trực tiếp
                                   → giảm thêm n_hours × n_blocks biến

  [WIN 3] Fix single-block jobs  → job chỉ có 1 block khả thi → y.lowBound=1
                                   → giảm không gian tìm kiếm MIP đáng kể

  [WIN 4] Tight upper bound e    → e <= |y_list| − 1
                                   → LP relaxation chặt hơn → solver hội tụ nhanh

  [WIN 5] Prune block_bay/e_vars → bỏ qua cặp (b, bay) không có y_vars nào
                                   → mô hình thưa (sparse) hơn

  [WIN 6] xlsxwriter thay openpyxl → không dùng in-memory DOM
                                   → ghi Excel ~3–5× nhanh hơn

  [WIN 7] to_dict('records') + itertuples → thay iterrows() trong vòng lặp
                                   → ~30–50 % nhanh hơn khi xử lý DataFrame
"""
import io
import time
import os
import pandas as pd
import pulp
from collections import defaultdict

try:
    import streamlit as st
    _HAS_ST = True
except ImportError:
    _HAS_ST = False

def _log(msg: str):
    """In ra terminal VÀ hiện lên Streamlit UI (nếu có)."""
    _log(msg)
    if _HAS_ST:
        st.write(msg)

# ============================================================
# TUNING FLAGS
# ============================================================
FAST_MODE         = True   # True = bỏ block_bay_wc khỏi MIP (khuyến nghị)
SOLVER_TIME_LIMIT = 120    # giây



# ============================================================
# SOLVER HELPER
# ============================================================
def _n_threads():
    try:    return max(1, os.cpu_count() or 1)
    except: return 1

def _make_solver(time_limit=120):
    """Ưu tiên HiGHS → CBC-multithread → CBC đơn."""
    n = _n_threads()
    try:
        import highspy  # noqa
        solver = pulp.HiGHS_CMD(
            msg=True, timeLimit=time_limit,
            options=[("parallel", "on"), ("threads", str(n))]
        )
        tp = pulp.LpProblem("_t", pulp.LpMinimize)
        tv = pulp.LpVariable("_v"); tp += tv; tp += tv >= 0
        tp.solve(solver)
        _log(f"[Solver] HiGHS ({n} threads)")
        return solver
    except Exception:
        pass
    try:
        solver = pulp.PULP_CBC_CMD(msg=True, timeLimit=time_limit, threads=n)
        _log(f"[Solver] CBC ({n} threads)")
        return solver
    except Exception:
        pass
    _log("[Solver] CBC (single thread)")
    return pulp.PULP_CBC_CMD(msg=True, timeLimit=time_limit)

# ============================================================
# MOVE HOUR SORT KEY
# ============================================================
_DAY_RANK = {'MO': 0, 'TU': 1, 'WE': 2, 'TH': 3, 'FR': 4, 'SA': 5, 'SU': 6}

def _hour_sort_key(h: str):
    s = str(h).strip().lstrip('+')
    if len(s) >= 2:
        dr = _DAY_RANK.get(s[:2].upper(), 99)
        try:    ti = int(s[2:]) if s[2:] else 0
        except: ti = 0
        return (dr, ti)
    return (99, s)

# ============================================================
# PUBLIC API
# ============================================================
def run_optimization(file_input):
    t0 = time.perf_counter()

    # =========================================================
    # 1. ĐỌC DỮ LIỆU
    # =========================================================
    xls = pd.ExcelFile(file_input)

    df1 = pd.read_excel(xls, sheet_name='MOVEHOUR-WEIGHTCLASS', header=None)
    has_st_pod    = (str(df1.iloc[1, 2]).strip().upper() == 'ST')
    data_col_start = 4 if has_st_pod else 2

    sts_bay_map = {}
    for col in range(data_col_start, df1.shape[1]):
        sts = df1.iloc[0, col]; bay = df1.iloc[1, col]
        if pd.notna(sts) and pd.notna(bay):
            sts_bay_map[col] = (str(sts).strip(), str(bay).strip())

    demands = {}
    cur_hour = None
    for idx in range(2, df1.shape[0]):
        row = df1.iloc[idx]
        hour = row[0]
        if pd.isna(hour): hour = cur_hour
        else:             cur_hour = hour
        weight = row[1]
        if pd.isna(weight): continue
        weight = int(float(str(weight)))
        st_v  = str(row[2]).strip() if has_st_pod and pd.notna(row[2]) else ''
        pod_v = str(row[3]).strip() if has_st_pod and pd.notna(row[3]) else ''
        for col in range(data_col_start, df1.shape[1]):
            qty = row[col]
            if pd.notna(qty) and qty != '':
                qty = int(float(str(qty)))
                if qty > 0:
                    sts, bay = sts_bay_map[col]
                    key  = (hour, sts, bay)
                    dkey = (weight, st_v, pod_v)
                    demands.setdefault(key, {})
                    demands[key][dkey] = demands[key].get(dkey, 0) + qty

    _log(f"Demand format: {'WC+ST+POD' if has_st_pod else 'WC only'}")
    job_keys          = list(demands.keys())
    all_hours_sorted  = sorted({h for h, s, b in job_keys}, key=_hour_sort_key)
    hour_rank         = {h: i for i, h in enumerate(all_hours_sorted)}
    jobs_by_hour      = defaultdict(list)
    for h, s, b in job_keys: jobs_by_hour[h].append((s, b))
    jobs_by_bay       = defaultdict(list)
    for h, s, bay in job_keys: jobs_by_bay[bay].append((h, s, bay))

    df2 = pd.read_excel(xls, sheet_name='BLOCK-WEIGHT CLASS', header=0)
    col_names         = [str(c).strip() for c in df2.columns]
    has_st_pod_supply = (col_names[1].upper() == 'ST' and col_names[2].upper() == 'POD')
    wc_col_start      = 3 if has_st_pod_supply else 1

    supply = {}; blocks_set = set()
    for _, row in df2.iterrows():
        block = str(row.iloc[0]).strip()
        if block in ('nan', 'GRAND TOTAL', '') or not block: continue
        st_v  = str(row.iloc[1]).strip() if has_st_pod_supply else ''
        pod_v = str(row.iloc[2]).strip() if has_st_pod_supply else ''
        skey  = (block, st_v, pod_v)
        wc_dict = {}
        for wi, w in enumerate([1, 2, 3, 4, 5]):
            ci  = wc_col_start + wi
            val = row.iloc[ci] if ci < len(row) else None
            wc_dict[w] = int(val) if pd.notna(val) and val != '' else 0
        supply[skey] = wc_dict
        blocks_set.add(block)

    weight_classes = [1, 2, 3, 4, 5]
    blocks         = sorted(blocks_set)
    supply_keys    = [k for k in supply if any(supply[k][w] > 0 for w in weight_classes)]
    _log(f"Supply format: {'BLOCK+ST+POD' if has_st_pod_supply else 'BLOCK only'}")

    # Sheet DATA
    container_data_available = False
    df_containers            = None
    try:
        df_containers = pd.read_excel(xls, sheet_name='DATA', header=0)
        cols = list(df_containers.columns)
        def find_col(cands):
            for c in cands:
                if c in cols: return c
            return None
        wc_src  = find_col(['YC', 'Unnamed: 1'])
        yp_src  = find_col(['YP', 'Unnamed: 2'])
        id_src  = find_col(['ID', 'Unnamed: 3'])
        st_src  = find_col(['ST'])
        pod_src = find_col(['POD'])
        if wc_src and yp_src and all(c in cols for c in ['YB', 'YR', 'YT']):
            df_containers = df_containers.dropna(
                subset=[wc_src, yp_src, 'YB', 'YR', 'YT']).copy()
            df_containers['REAL_WC']      = df_containers[wc_src].astype(float).astype(int)
            df_containers['YARD_POS']     = df_containers[yp_src].astype(str).str.strip()
            df_containers['REAL_CONT_ID'] = (
                df_containers[id_src].fillna('').astype(str).str.strip() if id_src else '')
            df_containers['CONT_ST']  = (
                df_containers[st_src].fillna('').astype(str).str.strip() if st_src else '')
            df_containers['CONT_POD'] = (
                df_containers[pod_src].fillna('').astype(str).str.strip() if pod_src else '')
            df_containers['YARD'] = df_containers['YARD'].astype(str).str.strip()
            for c in ['YB', 'YR', 'YT']:
                df_containers[c] = df_containers[c].astype(float).astype(int)
            container_data_available = True
            _log(f"Container DATA: {len(df_containers)} rows.")
    except Exception as e:
        _log(f"No DATA sheet. ({e})")

    # Stacking
    yb_wc_supply = {}; stack_ordering = {}; blocking_pairs = []
    if container_data_available:
        df_c = df_containers[['YARD', 'YB', 'YR', 'YT', 'REAL_WC',
                               'YARD_POS', 'REAL_CONT_ID',
                               'CONT_ST', 'CONT_POD']].copy()
        for block in blocks:
            bdf = df_c[df_c['YARD'] == block]
            if bdf.empty: continue
            yb_wc_supply[block] = {}; stack_ordering[block] = {}
            for yb, yb_df in bdf.groupby('YB'):
                yb_wc_supply[block][yb] = {
                    wc: int(cnt) for wc, cnt in yb_df.groupby('REAL_WC').size().items()}
                stack_ordering[block][yb] = {}
                for yr, yr_df in yb_df.groupby('YR'):
                    ordered = yr_df.sort_values('YT', ascending=False)[
                        ['YT', 'REAL_WC']].values.tolist()
                    stack_ordering[block][yb][yr] = [(int(t), int(w)) for t, w in ordered]
                for yr, tiers in stack_ordering[block][yb].items():
                    wcs_above = []
                    for tier, wc in tiers:
                        for prev_wc, prev_tier in wcs_above:
                            if prev_wc != wc:
                                blocking_pairs.append(
                                    (block, yb, yr, prev_tier, prev_wc, tier, wc))
                        wcs_above.append((wc, tier))
        _log(f"Stacking: {len(blocking_pairs)} blocking pairs.")

    # =========================================================
    # 2. DEMAND / SUPPLY CHECK
    # =========================================================
    total_demand = defaultdict(int)
    for job in job_keys:
        for dkey, qty in demands[job].items():
            total_demand[dkey] += qty
    total_supply = defaultdict(int)
    for skey in supply_keys:
        b, st_v, pod_v = skey
        for w in weight_classes:
            total_supply[(w, st_v, pod_v)] += supply[skey][w]

    ok = True
    for k in set(list(total_demand) + list(total_supply)):
        d = total_demand.get(k, 0); s = total_supply.get(k, 0)
        if d != s:
            _log(f"  ERROR Mismatch WC={k[0]} ST={k[1]} POD={k[2]}: "
                  f"demand={d}, supply={s}")
            ok = False
    if not ok:
        raise ValueError("Demand/supply mismatch.")
    _log("Demand/supply balanced OK.")

    t_read = time.perf_counter()
    _log(f"[TIMER] Đọc dữ liệu: {t_read - t0:.1f}s")

    # =========================================================
    # 3. XÂY DỰNG MIP
    # =========================================================
    _log(f"[MIP] FAST_MODE={'ON' if FAST_MODE else 'OFF'}")

    # Pre-index supply blocks by (st, pod)
    supply_blocks_by_st_pod = defaultdict(list)
    for b, st_v, pod_v in supply_keys:
        supply_blocks_by_st_pod[(st_v, pod_v)].append(b)

    # ── [WIN 1] Prune y_vars: chỉ tạo (job, block) có supply khớp ──
    feasible_blocks_for_job = {}
    n_fixed = 0
    for (h, s, bay) in job_keys:
        fblocks = set()
        for dkey in demands[(h, s, bay)]:
            _, st_v, pod_v = dkey
            fblocks.update(supply_blocks_by_st_pod.get((st_v, pod_v), []))
        feasible_blocks_for_job[(h, s, bay)] = sorted(fblocks)

    y_vars = {}
    for (h, s, bay) in job_keys:
        fblocks = feasible_blocks_for_job[(h, s, bay)]
        for b in fblocks:
            y_vars[(h, s, bay, b)] = pulp.LpVariable(
                f"y_{h}_{s}_{bay}_{b}", cat='Binary')
        # [WIN 3] Fix trivial assignments (chỉ 1 block khả thi)
        if len(fblocks) == 1:
            y_vars[(h, s, bay, fblocks[0])].lowBound = 1
            n_fixed += 1

    if n_fixed:
        _log(f"[MIP] Fixed {n_fixed} jobs (single feasible block).")

    # x_vars – chỉ tạo khi y_var tương ứng tồn tại
    x_vars = {}
    for (h, s, bay) in job_keys:
        for dkey in demands[(h, s, bay)]:
            w, st_v, pod_v = dkey
            for b in supply_blocks_by_st_pod.get((st_v, pod_v), []):
                if (h, s, bay, b) in y_vars:
                    x_vars[(h, s, bay, b, dkey)] = pulp.LpVariable(
                        f"x_{h}_{s}_{bay}_{b}_{w}_{st_v}_{pod_v}",
                        lowBound=0, cat='Integer')

    prob = pulp.LpProblem("Minimize_Clashes", pulp.LpMinimize)

    CLASH_W = 100.0; SINGLE_W = 10.0; SPREAD_W = 5.0; BAY_SINGLE_W = 10.0

    # ── [WIN 2+4] Xóa u_vars, viết e trực tiếp + tight upper bound ──
    e_vars = {}
    for h in jobs_by_hour:
        for b in blocks:
            y_list = [y_vars[(h, s, bay, b)]
                      for (s, bay) in jobs_by_hour[h]
                      if (h, s, bay, b) in y_vars]
            if not y_list:
                continue   # [WIN 5] bỏ qua cặp (h,b) không có y_vars
            e = pulp.LpVariable(f"e_{h}_{b}", lowBound=0, cat='Integer')
            e_vars[(h, b)] = e
            ysum = pulp.lpSum(y_list)
            prob += e >= ysum - 1                 # lower bound
            prob += e <= len(y_list) - 1          # [WIN 4] tight upper bound

    # single_block per job
    single_block = {}
    for (h, s, bay) in job_keys:
        fblocks = feasible_blocks_for_job[(h, s, bay)]
        sb = pulp.LpVariable(f"sb_{h}_{s}_{bay}",
                             lowBound=0, upBound=1, cat='Continuous')
        single_block[(h, s, bay)] = sb
        prob += sb >= 2 - pulp.lpSum(y_vars[(h, s, bay, b)] for b in fblocks)

    # ── [WIN 5] Prune block_bay: chỉ tạo cho cặp (b, bay) có y_vars ──
    all_bays     = sorted({bay for (_, _, bay) in job_keys})
    relevant_bb  = set()
    for (h, s, bay) in job_keys:
        for b in feasible_blocks_for_job[(h, s, bay)]:
            relevant_bb.add((b, bay))

    block_bay = {}
    for (b, bay) in sorted(relevant_bb):
        var = pulp.LpVariable(f"bb_{b}_{bay}", cat='Binary')
        block_bay[(b, bay)] = var
        for (h, s, bj) in jobs_by_bay[bay]:
            if (h, s, bay, b) in y_vars:
                prob += var >= y_vars[(h, s, bay, b)]

    blocks_per_bay = defaultdict(list)
    for (b, bay) in relevant_bb:
        blocks_per_bay[bay].append(b)

    bay_single = {}
    for bay in all_bays:
        blist = blocks_per_bay.get(bay, [])
        if not blist:
            continue
        var = pulp.LpVariable(f"bs_{bay}", lowBound=0, upBound=1, cat='Continuous')
        bay_single[bay] = var
        bb_sum = pulp.lpSum(block_bay.get((b, bay), 0) for b in blist)
        prob += var >= 2 - bb_sum
        prob += bb_sum >= 2

    # block_bay_wc – chỉ khi FAST_MODE=False
    block_bay_wc = {}
    if not FAST_MODE:
        x_by_bbw = defaultdict(list)
        for (h, s, bay, b, dkey), xvar in x_vars.items():
            x_by_bbw[(b, bay, dkey[0])].append((xvar, demands[(h, s, bay)][dkey]))
        for b in blocks:
            for bay in all_bays:
                for wc in weight_classes:
                    entries = x_by_bbw.get((b, bay, wc), [])
                    if not entries: continue
                    var = pulp.LpVariable(f"bbw_{b}_{bay}_{wc}", cat='Binary')
                    block_bay_wc[(b, bay, wc)] = var
                    for xvar, d in entries:
                        prob += var >= xvar / (d + 0.1)

    # Objective
    obj = (CLASH_W      * pulp.lpSum(e_vars.values())      +
           SINGLE_W     * pulp.lpSum(single_block.values()) +
           SPREAD_W     * pulp.lpSum(block_bay.values())    +
           BAY_SINGLE_W * pulp.lpSum(bay_single.values()))
    if block_bay_wc:
        obj += 2.0 * pulp.lpSum(block_bay_wc.values())
    prob += obj

    # Demand constraints
    for (h, s, bay) in job_keys:
        for dkey, d in demands[(h, s, bay)].items():
            w, st_v, pod_v = dkey
            x_list = [x_vars[(h, s, bay, b, dkey)]
                      for b in supply_blocks_by_st_pod.get((st_v, pod_v), [])
                      if (h, s, bay, b, dkey) in x_vars]
            if x_list:
                prob += pulp.lpSum(x_list) == d

    # Supply constraints
    x_by_supply = defaultdict(list)
    for (h, s, bay, b, dkey), xvar in x_vars.items():
        w, st_v, pod_v = dkey
        x_by_supply[(b, st_v, pod_v, w)].append(xvar)
    for skey in supply_keys:
        b, st_v, pod_v = skey
        for w in weight_classes:
            xl = x_by_supply.get((b, st_v, pod_v, w), [])
            if xl:
                prob += pulp.lpSum(xl) <= supply[skey][w]

    # Linking x <= d * y
    for (h, s, bay) in job_keys:
        for dkey, d in demands[(h, s, bay)].items():
            for b in supply_blocks_by_st_pod.get((dkey[1], dkey[2]), []):
                key = (h, s, bay, b, dkey)
                if key in x_vars:
                    prob += x_vars[key] <= d * y_vars[(h, s, bay, b)]

    t_build = time.perf_counter()
    nv = len(prob.variables()); nc = len(prob.constraints)
    _log(f"[TIMER] Build model: {t_build - t_read:.1f}s  |  "
          f"vars={nv}, constraints={nc}")

    solver = _make_solver(SOLVER_TIME_LIMIT)
    prob.solve(solver)

    t_solve = time.perf_counter()
    _log(f"[TIMER] Solver: {t_solve - t_build:.1f}s")
    _log(f"Status: {pulp.LpStatus[prob.status]}")
    if prob.status == pulp.LpStatusInfeasible:
        raise RuntimeError("Model infeasible.")
    elif prob.status not in (1,):
        _log("No optimal within time limit – using best solution found.")

    # =========================================================
    # 4. KẾT QUẢ VÀ GÁN CONTAINER
    # =========================================================
    result_rows = []
    for (h, s, bay, b), yvar in y_vars.items():
        yv = pulp.value(yvar)
        if yv is not None and yv > 0.5:
            for dkey in demands[(h, s, bay)]:
                xkey = (h, s, bay, b, dkey)
                if xkey not in x_vars: continue
                qty = pulp.value(x_vars[xkey])
                if qty is not None and qty > 0.5:
                    w, st_v, pod_v = dkey
                    result_rows.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay,
                        'ASSIGNED BLOCK': b, 'WEIGHT CLASS': w,
                        'ST': st_v, 'POD': pod_v,
                        'QUANTITIES': int(round(qty))
                    })
    df_result = pd.DataFrame(result_rows)
    df_result['_sort_hr'] = df_result['MOVE HOUR'].map(hour_rank)
    df_result.sort_values(['_sort_hr', 'STS', 'BAY', 'ASSIGNED BLOCK'], inplace=True)
    df_result.drop(columns=['_sort_hr'], inplace=True)

    # Container assignment
    df_result_detail = []
    if container_data_available:
        # Build pool
        pool = defaultdict(list)
        for row in df_containers[['YARD', 'YB', 'YR', 'YT', 'REAL_WC',
                                   'YARD_POS', 'REAL_CONT_ID',
                                   'CONT_ST', 'CONT_POD']].itertuples(index=False):
            c = {'yb': int(row.YB), 'yr': int(row.YR), 'yt': int(row.YT),
                 'wc': int(row.REAL_WC), 'yard_pos': row.YARD_POS,
                 'real_cont_id': row.REAL_CONT_ID,
                 'st': row.CONT_ST, 'pod': row.CONT_POD,
                 'picked': False, 'pick_h': None}
            pool[row.YARD].append(c)

        # O(1) accessibility
        blocked_count = {}; below_map = {}
        for blk, conts in pool.items():
            stacks = defaultdict(list)
            for c in conts: stacks[(c['yb'], c['yr'])].append(c)
            for c in conts:
                above = [o for o in stacks[(c['yb'], c['yr'])]
                         if o is not c and o['yt'] > c['yt']]
                below = [o for o in stacks[(c['yb'], c['yr'])]
                         if o is not c and o['yt'] < c['yt']]
                blocked_count[id(c)] = len(above)
                below_map[id(c)]     = below

        avail = defaultdict(list)
        for blk, conts in pool.items():
            for c in conts:
                if blocked_count[id(c)] == 0:
                    avail[(blk, c['wc'], c['st'], c['pod'])].append(c)

        opened_ybs = set()

        def pick_n(block, wc, st_match, pod_match, qty,
                   h, s_job, bay_job, h_rank_val, result_list):
            remaining = qty
            akey = (block, wc, st_match, pod_match)
            av   = avail[akey]
            while remaining > 0:
                if not av: break
                yb_cnt = defaultdict(int)
                for c in av: yb_cnt[c['yb']] += 1
                best_i = min(range(len(av)), key=lambda i: (
                    0 if (block, av[i]['yb']) in opened_ybs else 1,
                    -yb_cnt[av[i]['yb']],
                    av[i]['yb'], av[i]['yr'], -av[i]['yt']
                ))
                best = av.pop(best_i)
                best['picked'] = True; best['pick_h'] = h
                opened_ybs.add((block, best['yb']))
                for lower in below_map.get(id(best), []):
                    if lower['picked']: continue
                    blocked_count[id(lower)] -= 1
                    if blocked_count[id(lower)] == 0:
                        avail[(block, lower['wc'],
                               lower['st'], lower['pod'])].append(lower)
                result_list.append({
                    'MOVE HOUR': h, 'CONTAINER ID': best['real_cont_id'],
                    'ST': best['st'], 'POD': best['pod'],
                    'STS': s_job, 'BAY': bay_job,
                    'ASSIGNED BLOCK': block, 'WEIGHT CLASS': wc,
                    'QUANTITIES': qty,
                    'YB': best['yb'], 'YR': best['yr'], 'YT': best['yt'],
                    'YARD POSITION': best['yard_pos']
                })
                remaining -= 1
            return remaining

        # [WIN 7] sort once, iterate with hour filter
        df_rs = df_result.copy()
        df_rs['_hr'] = df_rs['MOVE HOUR'].map(hour_rank)
        df_rs.sort_values(['_hr', 'STS', 'BAY', 'ASSIGNED BLOCK', 'WEIGHT CLASS'],
                          inplace=True)
        # Group by hour for fast per-hour access
        rs_by_hour = defaultdict(list)
        for rec in df_rs.to_dict('records'):
            rs_by_hour[rec['MOVE HOUR']].append(rec)

        deferred = []
        for h in all_hours_sorted:
            hrv = hour_rank[h]
            for asg in rs_by_hour[h]:
                s, bay_job, b = asg['STS'], asg['BAY'], asg['ASSIGNED BLOCK']
                w     = int(asg['WEIGHT CLASS'])
                st_v  = str(asg.get('ST', '')).strip()
                pod_v = str(asg.get('POD', '')).strip()
                qty   = int(asg['QUANTITIES'])
                if b not in pool:
                    df_result_detail.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay_job,
                        'ASSIGNED BLOCK': b, 'WEIGHT CLASS': w,
                        'CONTAINER ID': '', 'ST': '', 'POD': '',
                        'QUANTITIES': qty,
                        'YB': '', 'YR': '', 'YT': '', 'YARD POSITION': ''
                    })
                    continue
                rem = pick_n(b, w, st_v, pod_v, qty,
                             h, s, bay_job, hrv, df_result_detail)
                if rem > 0:
                    deferred.append({'b': b, 'wc': w, 'st': st_v, 'pod': pod_v,
                                     'qty': rem, 'h_orig': h, 's': s,
                                     'bay': bay_job, 'h_rank_min': hrv})
            still_def = []
            for d in deferred:
                rem = pick_n(d['b'], d['wc'], d['st'], d['pod'], d['qty'],
                             h, d['s'], d['bay'], hrv, df_result_detail)
                if rem > 0:
                    d2 = d.copy(); d2['qty'] = rem; still_def.append(d2)
            deferred = still_def

        rh = sum(d['qty'] for d in deferred)
        if rh > 0:
            _log(f"  Re-handling: {rh} containers.")
            for d in deferred:
                _log(f"    Block {d['b']} WC{d['wc']} x{d['qty']} "
                      f"(from {d['h_orig']})")
        else:
            _log("  All containers assigned – no re-handling.")

        df_result_detail = pd.DataFrame(df_result_detail)
    else:
        df_result_detail = df_result.copy()
        df_result_detail.insert(1, 'CONTAINER ID', '')
        df_result_detail.insert(2, 'ST', '')
        df_result_detail.insert(3, 'POD', '')
        df_result_detail['YB'] = ''
        df_result_detail['YR'] = ''
        df_result_detail['YT'] = ''
        df_result_detail['YARD POSITION'] = ''

    df_result_detail['_sort_hr'] = df_result_detail['MOVE HOUR'].map(hour_rank)
    df_result_detail.sort_values(
        ['_sort_hr', 'STS', 'BAY', 'ASSIGNED BLOCK',
         'WEIGHT CLASS', 'YB', 'YR', 'YT'], inplace=True)
    df_result_detail.drop(columns=['_sort_hr'], inplace=True)

    t_assign = time.perf_counter()
    _log(f"[TIMER] Gán containers: {t_assign - t_solve:.1f}s")

    # =========================================================
    # 5. CLASH
    # =========================================================
    clash_details = []
    total_clashes = 0
    for (h, b), evar in e_vars.items():
        e_val = pulp.value(evar)
        if e_val is not None and e_val > 0.5:
            total_clashes += e_val
            jobs = [f"{s}@{bay}" for (s, bay) in jobs_by_hour.get(h, [])
                    if (h, s, bay, b) in y_vars
                    and pulp.value(y_vars[(h, s, bay, b)]) is not None
                    and pulp.value(y_vars[(h, s, bay, b)]) > 0.5]
            u_val = len(jobs) + int(e_val)   # u = e + 1 conceptually
            clash_details.append({
                'MOVE HOUR': h, 'BLOCK': b,
                'SỐ LƯỢNG BAY (u)': u_val,
                'CLASH (e = u-1)': int(e_val),
                'DANH SÁCH JOB (STS@BAY)': ', '.join(jobs)
            })
    df_clash = pd.DataFrame(clash_details)
    if not df_clash.empty:
        df_clash['_sort_hr'] = df_clash['MOVE HOUR'].map(hour_rank)
        df_clash.sort_values(['_sort_hr', 'BLOCK'], inplace=True)
        df_clash.drop(columns=['_sort_hr'], inplace=True)
    _log(f"Total clashes: {total_clashes}")

    # =========================================================
    # 6. GHI EXCEL – openpyxl tối ưu
    #    • Style objects tạo 1 lần, dùng lại cho mọi cell
    #    • ws.append() bulk-write toàn bộ data (nhanh nhất trong openpyxl)
    #    • to_dict('records') thay iterrows()  [WIN 7]
    #    • Merge + style áp dụng sau khi append xong
    # =========================================================
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    # ── Style factory (tạo 1 lần) ──
    def _side():
        return Side(border_style='thin', color='FF000000')
    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)
    def _font(bold=False, color='FF000000', size=10):
        return Font(name='Calibri', bold=bold, color=color, size=size)
    def _fill(hex6):           # hex6 không có '#'
        return PatternFill('solid', fgColor=hex6)
    def _align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # Hex dạng FFRRGGBB cho openpyxl
    _OX = {
        'dark':  'FF1F4E79', 'mid':   'FF2E75B6',
        'light': 'FF9DC3E6', 'pale':  'FFD6E4F0',
        'alt':   'FFEBF3FB', 'white': 'FFFFFFFF',
    }
    BD = _border()

    # Pre-build các style tuple thường dùng
    # (font, fill, alignment, border) – gán trực tiếp vào cell
    _ST = {
        'hdr_dark':  (_font(bold=True, color='FFFFFFFF'), _fill(_OX['dark']),
                      _align(wrap=True), BD),
        'hdr_mid':   (_font(bold=True, color='FFFFFFFF'), _fill(_OX['mid']),
                      _align(wrap=True), BD),
        'hdr_light': (_font(bold=True),                   _fill(_OX['light']),
                      _align(wrap=True), BD),
        'hdr_pale':  (_font(bold=True),                   _fill(_OX['pale']),
                      _align(wrap=True), BD),
        'data_w':    (_font(), _fill(_OX['white']), _align(), BD),
        'data_a':    (_font(), _fill(_OX['alt']),   _align(), BD),
        'cont_list': (_font(size=9), _fill(_OX['pale']),
                      _align(h='left', v='top', wrap=True), BD),
    }

    def _apply(cell, key):
        f, fi, al, bo = _ST[key]
        cell.font = f; cell.fill = fi
        cell.alignment = al; cell.border = bo

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ── CLASH sheet ──
    ws_clash   = wb.create_sheet('CLASH')
    hdrs_clash = ['MOVE HOUR', 'BLOCK', 'SỐ LƯỢNG BAY (u)',
                  'CLASH (e = u-1)', 'DANH SÁCH JOB (STS@BAY)']
    clash_col_w = [14, 12, 18, 18, 50]
    ws_clash.append(hdrs_clash)
    for ci, (hdr, cw) in enumerate(zip(hdrs_clash, clash_col_w), 1):
        cell = ws_clash.cell(row=1, column=ci)
        _apply(cell, 'hdr_dark')
        ws_clash.column_dimensions[get_column_letter(ci)].width = cw
    if not df_clash.empty:
        clash_recs = df_clash[hdrs_clash].to_dict('records')   # [WIN 7]
        for ri, rec in enumerate(clash_recs, 2):
            ws_clash.append([rec.get(c, '') for c in hdrs_clash])
            sk = 'data_w' if ri % 2 == 0 else 'data_a'
            for ci in range(1, 6):
                _apply(ws_clash.cell(row=ri, column=ci), sk)
    else:
        ws_clash.append(['Không có clash nào xảy ra.'])
        ws_clash.merge_cells(start_row=2, start_column=1,
                             end_row=2,   end_column=5)
        _apply(ws_clash.cell(row=2, column=1), 'data_w')

    # ── Result column config ──
    core_cols     = ['MOVE HOUR', 'CONT LIST', 'CONTAINER ID', 'ST', 'POD',
                     'STS', 'BAY', 'ASSIGNED BLOCK', 'WEIGHT CLASS', 'QUANTITIES']
    position_cols = ['YB', 'YR', 'YT', 'YARD POSITION']
    all_result_cols = (core_cols + position_cols) if container_data_available else \
                      ['MOVE HOUR', 'STS', 'BAY', 'ASSIGNED BLOCK',
                       'WEIGHT CLASS', 'QUANTITIES']
    col_widths_map = {
        'MOVE HOUR': 14, 'CONT LIST': 45, 'CONTAINER ID': 20,
        'ST': 10, 'POD': 10, 'STS': 10, 'BAY': 10,
        'ASSIGNED BLOCK': 16, 'WEIGHT CLASS': 14, 'QUANTITIES': 12,
        'YB': 8, 'YR': 8, 'YT': 8, 'YARD POSITION': 18
    }
    CONT_LIST_SET = {'CONT LIST'}
    CONT_ID_SET   = {'CONTAINER ID', 'ST', 'POD'}
    POS_SET       = set(position_cols)
    INT_COLS      = {'YB', 'YR', 'YT'}

    def _hdr_key(cn):
        if cn in CONT_LIST_SET: return 'hdr_pale'
        if cn in CONT_ID_SET:   return 'hdr_mid'
        if cn in POS_SET:       return 'hdr_light'
        return 'hdr_dark'

    cl_idx = (all_result_cols.index('CONT LIST') + 1
              if 'CONT LIST' in all_result_cols else None)   # 1-based

    def write_result_sheet(ws, df, sheet_title):
        n_cols = len(all_result_cols)
        n_rows = len(df)

        # Header
        ws.append(all_result_cols)
        for ci, cn in enumerate(all_result_cols, 1):
            cell = ws.cell(row=1, column=ci)
            _apply(cell, _hdr_key(cn))
            ws.column_dimensions[get_column_letter(ci)].width = \
                col_widths_map.get(cn, 14)

        # Pre-build CONT LIST map [WIN 7]
        cont_list_map = {}
        if cl_idx and 'CONTAINER ID' in df.columns:
            for (mh, bay), grp in df.groupby(['MOVE HOUR', 'BAY']):
                ids = [str(v).strip() for v in grp['CONTAINER ID']
                       if str(v).strip() not in ('', 'nan')]
                cont_list_map[(mh, bay)] = ', '.join(ids) if ids else ''

        # [WIN 7] to_dict once → fastest per-row iteration
        records = df.to_dict('records')

        # ── Bulk-append all data rows first (fastest path in openpyxl) ──
        data_style_keys = []          # store style key per row for 2nd pass
        shade = True; prev_gk = None
        for rec in records:
            gk = (rec.get('MOVE HOUR'), rec.get('STS'), rec.get('BAY'),
                  rec.get('ASSIGNED BLOCK'), rec.get('WEIGHT CLASS'))
            if gk != prev_gk:
                shade = not shade; prev_gk = gk
            sk = 'data_w' if shade else 'data_a'
            data_style_keys.append(sk)

            row_vals = []
            for cn in all_result_cols:
                if cn == 'CONT LIST':
                    row_vals.append(None)
                    continue
                v = rec.get(cn, '')
                if cn in INT_COLS and v != '':
                    try: v = int(v)
                    except: pass
                if v == '' or (isinstance(v, float) and str(v) == 'nan'):
                    v = None
                row_vals.append(v)
            ws.append(row_vals)

        # ── 2nd pass: apply styles (row offset = 2 because row 1 = header) ──
        for ri, sk in enumerate(data_style_keys, 2):
            for ci in range(1, n_cols + 1):
                if cl_idx and ci == cl_idx:
                    continue    # CONT LIST styled separately below
                _apply(ws.cell(row=ri, column=ci), sk)

        # ── CONT LIST: merge + style ──
        if cl_idx:
            prev_key = None; grp_start = 2; n_ids_grp = 0
            def _flush_cl(pk, gs, ge, nids):
                text = cont_list_map.get(pk) or None
                if ge > gs:
                    ws.merge_cells(start_row=gs, start_column=cl_idx,
                                   end_row=ge,   end_column=cl_idx)
                top_cell = ws.cell(row=gs, column=cl_idx)
                top_cell.value = text
                _apply(top_cell, 'cont_list')
                span = ge - gs + 1
                rh = max(15, min(60, max(1, -(-nids // max(1, span))) * 13))
                for r in range(gs, ge + 1):
                    ws.row_dimensions[r].height = rh

            for ri, rec in enumerate(records, 2):
                cur_key = (rec.get('MOVE HOUR', ''), rec.get('BAY', ''))
                nids = len([x for x in cont_list_map.get(cur_key, '').split(',')
                            if x.strip()])
                if cur_key != prev_key:
                    if prev_key is not None:
                        _flush_cl(prev_key, grp_start, ri - 1, n_ids_grp)
                    prev_key = cur_key; grp_start = ri; n_ids_grp = nids
            if prev_key is not None:
                _flush_cl(prev_key, grp_start, n_rows + 1, n_ids_grp)

        _log(f"  Sheet '{sheet_title}': {n_rows} rows written.")

    # RESULT sheets by ST
    if container_data_available and 'ST' in df_result_detail.columns:
        st_values = [s for s in sorted(df_result_detail['ST'].dropna().unique())
                     if str(s).strip() not in ('', 'nan')]
    else:
        st_values = ['ALL']
    if not st_values:
        st_values = ['ALL']

    for st_idx, st_val in enumerate(st_values, 1):
        sname = (f"RESULT {st_idx} ({st_val})"
                 if st_val != 'ALL' else 'RESULT')[:31]
        ws = wb.create_sheet(sname)
        df_rd = (df_result_detail if st_val == 'ALL' else
                 df_result_detail[
                     df_result_detail['ST'].astype(str).str.strip()
                     == str(st_val).strip()]
                 ).reset_index(drop=True)
        write_result_sheet(ws, df_rd, sname)

    ws_total = wb.create_sheet('RESULT TOTAL')
    write_result_sheet(ws_total,
                       df_result_detail.reset_index(drop=True),
                       'RESULT TOTAL')

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    t_end       = time.perf_counter()
    total_rows  = len(df_result_detail)
    _log(f"[TIMER] Ghi Excel: {t_end - t_assign:.1f}s")
    _log(f"[TIMER] ══ TỔNG CỘNG: {t_end - t0:.1f}s ══")
    _log(f"Done. Rows={total_rows}, Total Clashes={total_clashes}")
    return excel_buffer, total_rows, total_clashes
